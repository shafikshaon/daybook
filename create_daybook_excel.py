#!/usr/bin/env python3
"""
Daybook Financial Tracker - Excel Template Generator
Creates a ready-to-use Excel file with all sheets, formulas, and sample data
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import random

def create_daybook_excel():
    """Create the complete Daybook Excel template"""

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create all sheets
    sheets = [
        'Dashboard', 'Profile', 'Accounts', 'Transactions', 'Recurring',
        'Tags', 'Budgets', 'Goals', 'Credit_Cards', 'CC_Transactions',
        'CC_Payments', 'Bills', 'Bill_Payments', 'Reports', 'Charts'
    ]

    for sheet_name in sheets:
        wb.create_sheet(sheet_name)

    # Styling
    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== 1. DASHBOARD =====
    ws = wb['Dashboard']

    # Headers
    ws['A1'] = 'Metric'
    ws['B1'] = 'Value'
    ws['C1'] = 'Change'
    ws['D1'] = 'Status'

    # Apply header style
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].font = header_font
        ws[f'{col}1'].alignment = Alignment(horizontal='center')

    # Metrics
    metrics = [
        ('Total Net Worth', '=SUM(Accounts!D:D)', '', ''),
        ('Total Income (This Month)', '=SUMIFS(Transactions!C:C,Transactions!B:B,"Income",Transactions!A:A,">="&EOMONTH(TODAY(),-1)+1,Transactions!A:A,"<="&EOMONTH(TODAY(),0))', '', ''),
        ('Total Expenses (This Month)', '=SUMIFS(Transactions!C:C,Transactions!B:B,"Expense",Transactions!A:A,">="&EOMONTH(TODAY(),-1)+1,Transactions!A:A,"<="&EOMONTH(TODAY(),0))', '', ''),
        ('Net Savings (This Month)', '=B3-B4', '', ''),
        ('Budget Progress', '=B4/SUM(Budgets!C:C)', '', ''),
        ('Active Accounts', '=COUNTA(Accounts!B:B)-1', '', ''),
        ('Active Credit Cards', '=COUNTA(Credit_Cards!B:B)-1', '', ''),
        ('Pending Bills', '=COUNTIFS(Bills!F:F,"Pending")', '', ''),
        ('Goals Progress', '=AVERAGE(Goals!E:E)', '', ''),
    ]

    for idx, metric in enumerate(metrics, start=2):
        ws[f'A{idx}'] = metric[0]
        ws[f'B{idx}'] = metric[1]
        ws[f'C{idx}'] = metric[2]
        ws[f'D{idx}'] = metric[3]

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

    # ===== 2. PROFILE =====
    ws = wb['Profile']
    ws['A1'] = 'Setting'
    ws['B1'] = 'Value'

    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['B1'].fill = header_fill
    ws['B1'].font = header_font

    profile_data = [
        ('Name', 'Your Name'),
        ('Currency', 'USD'),
        ('Date Format', 'MM/DD/YYYY'),
        ('Start Date', '01/01/2024'),
        ('Fiscal Year Start', 'January'),
        ('Budget Period', 'Monthly'),
    ]

    for idx, data in enumerate(profile_data, start=2):
        ws[f'A{idx}'] = data[0]
        ws[f'B{idx}'] = data[1]

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20

    # ===== 3. ACCOUNTS =====
    ws = wb['Accounts']
    headers = ['ID', 'Account Name', 'Type', 'Balance', 'Last Updated', 'Status', 'Notes']

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Sample accounts
    accounts = [
        (1, 'Chase Checking', 'Checking', 5000.00, datetime.now(), 'Active', 'Primary checking'),
        (2, 'Savings Account', 'Savings', 15000.00, datetime.now(), 'Active', 'Emergency fund'),
        (3, 'Investment Account', 'Investment', 50000.00, datetime.now(), 'Active', 'Retirement'),
        (4, 'Cash Wallet', 'Cash', 200.00, datetime.now(), 'Active', 'Daily cash'),
    ]

    for row_idx, account in enumerate(accounts, start=2):
        for col_idx, value in enumerate(account, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Total row
    total_row = len(accounts) + 2
    ws[f'C{total_row}'] = 'TOTAL'
    ws[f'C{total_row}'].font = Font(bold=True)
    ws[f'D{total_row}'] = f'=SUM(D2:D{total_row-1})'
    ws[f'D{total_row}'].font = Font(bold=True)

    # Data validation for Type
    account_types = DataValidation(type="list", formula1='"Checking,Savings,Investment,Cash,Credit Card,Loan,Other"', allow_blank=False)
    ws.add_data_validation(account_types)
    account_types.add(f'C2:C100')

    # Data validation for Status
    status_validation = DataValidation(type="list", formula1='"Active,Inactive,Closed"', allow_blank=False)
    ws.add_data_validation(status_validation)
    status_validation.add(f'F2:F100')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 25

    # ===== 4. TRANSACTIONS =====
    ws = wb['Transactions']
    headers = ['Date', 'Type', 'Amount', 'Category', 'Account', 'Description', 'Tags', 'Recurring', 'Status']

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Sample transactions
    base_date = datetime(2024, 1, 1)
    transactions = [
        (base_date + timedelta(days=14), 'Income', 5000.00, 'Salary', 'Chase Checking', 'Monthly salary', 'Income:Salary', 'No', 'Completed'),
        (base_date + timedelta(days=15), 'Expense', 1200.00, 'Rent', 'Chase Checking', 'Apartment rent', 'Housing:Rent', 'Yes', 'Completed'),
        (base_date + timedelta(days=16), 'Expense', 150.00, 'Groceries', 'Chase Checking', 'Weekly groceries', 'Food:Groceries', 'No', 'Completed'),
        (base_date + timedelta(days=17), 'Expense', 50.00, 'Gas', 'Chase Checking', 'Car fuel', 'Transportation:Gas', 'No', 'Completed'),
        (base_date + timedelta(days=18), 'Expense', 75.00, 'Dining Out', 'Chase Checking', 'Restaurant', 'Food:Dining', 'No', 'Completed'),
        (base_date + timedelta(days=19), 'Expense', 15.99, 'Subscription', 'Chase Checking', 'Netflix', 'Entertainment:Streaming', 'Yes', 'Completed'),
        (base_date + timedelta(days=20), 'Expense', 45.00, 'Utilities', 'Chase Checking', 'Electric bill', 'Housing:Utilities', 'Yes', 'Completed'),
        (base_date + timedelta(days=21), 'Expense', 120.00, 'Shopping', 'Chase Checking', 'Clothing', 'Shopping:Clothing', 'No', 'Completed'),
        (base_date + timedelta(days=22), 'Expense', 30.00, 'Entertainment', 'Cash Wallet', 'Movie tickets', 'Entertainment:Movies', 'No', 'Completed'),
        (base_date + timedelta(days=23), 'Expense', 200.00, 'Groceries', 'Chase Checking', 'Costco run', 'Food:Groceries', 'No', 'Completed'),
    ]

    for row_idx, transaction in enumerate(transactions, start=2):
        for col_idx, value in enumerate(transaction, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Data validations
    type_validation = DataValidation(type="list", formula1='"Income,Expense,Transfer"', allow_blank=False)
    ws.add_data_validation(type_validation)
    type_validation.add('B2:B1000')

    recurring_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(recurring_validation)
    recurring_validation.add('H2:H1000')

    status_validation = DataValidation(type="list", formula1='"Completed,Pending,Cancelled"', allow_blank=False)
    ws.add_data_validation(status_validation)
    status_validation.add('I2:I1000')

    # Summary section
    ws['M1'] = 'THIS MONTH SUMMARY'
    ws['M1'].font = Font(bold=True, size=12)
    ws['M2'] = 'Total Income'
    ws['N2'] = '=SUMIFS(C:C,B:B,"Income",A:A,">="&EOMONTH(TODAY(),-1)+1,A:A,"<="&EOMONTH(TODAY(),0))'
    ws['M3'] = 'Total Expenses'
    ws['N3'] = '=SUMIFS(C:C,B:B,"Expense",A:A,">="&EOMONTH(TODAY(),-1)+1,A:A,"<="&EOMONTH(TODAY(),0))'
    ws['M4'] = 'Net Savings'
    ws['N4'] = '=N2-N3'
    ws['M5'] = 'Transaction Count'
    ws['N5'] = '=COUNTIFS(A:A,">="&EOMONTH(TODAY(),-1)+1,A:A,"<="&EOMONTH(TODAY(),0))'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12

    # ===== 5. RECURRING =====
    ws = wb['Recurring']
    headers = ['ID', 'Name', 'Type', 'Amount', 'Category', 'Account', 'Frequency', 'Start Date', 'End Date', 'Active']

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    recurring = [
        (1, 'Monthly Salary', 'Income', 5000.00, 'Salary', 'Chase Checking', 'Monthly', datetime(2024, 1, 1), None, 'Yes'),
        (2, 'Rent Payment', 'Expense', 1200.00, 'Rent', 'Chase Checking', 'Monthly', datetime(2024, 1, 1), None, 'Yes'),
        (3, 'Netflix', 'Expense', 15.99, 'Subscription', 'Chase Checking', 'Monthly', datetime(2024, 1, 1), None, 'Yes'),
        (4, 'Gym Membership', 'Expense', 50.00, 'Health', 'Chase Checking', 'Monthly', datetime(2024, 1, 1), None, 'Yes'),
    ]

    for row_idx, item in enumerate(recurring, start=2):
        for col_idx, value in enumerate(item, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Data validations
    type_validation = DataValidation(type="list", formula1='"Income,Expense"', allow_blank=False)
    ws.add_data_validation(type_validation)
    type_validation.add('C2:C100')

    frequency_validation = DataValidation(type="list", formula1='"Daily,Weekly,Monthly,Yearly"', allow_blank=False)
    ws.add_data_validation(frequency_validation)
    frequency_validation.add('G2:G100')

    active_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(active_validation)
    active_validation.add('J2:J100')

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # ===== 6. TAGS =====
    ws = wb['Tags']
    headers = ['ID', 'Tag Name', 'Type', 'Parent', 'Color', 'Icon']

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    tags = [
        (1, 'Income', 'Income', '', '#4CAF50', '💰'),
        (2, 'Salary', 'Income', 'Income', '#4CAF50', '💵'),
        (3, 'Freelance', 'Income', 'Income', '#4CAF50', '💼'),
        (4, 'Expenses', 'Expense', '', '#F44336', '💸'),
        (5, 'Housing', 'Expense', 'Expenses', '#F44336', '🏠'),
        (6, 'Rent', 'Expense', 'Housing', '#F44336', '🏠'),
        (7, 'Utilities', 'Expense', 'Housing', '#F44336', '⚡'),
        (8, 'Food', 'Expense', 'Expenses', '#FF9800', '🍔'),
        (9, 'Groceries', 'Expense', 'Food', '#FF9800', '🛒'),
        (10, 'Dining Out', 'Expense', 'Food', '#FF9800', '🍽️'),
        (11, 'Transportation', 'Expense', 'Expenses', '#2196F3', '🚗'),
        (12, 'Gas', 'Expense', 'Transportation', '#2196F3', '⛽'),
        (13, 'Public Transit', 'Expense', 'Transportation', '#2196F3', '🚌'),
        (14, 'Entertainment', 'Expense', 'Expenses', '#9C27B0', '🎬'),
        (15, 'Shopping', 'Expense', 'Expenses', '#E91E63', '🛍️'),
        (16, 'Health', 'Expense', 'Expenses', '#00BCD4', '🏥'),
        (17, 'Subscription', 'Expense', 'Expenses', '#795548', '📱'),
    ]

    for row_idx, tag in enumerate(tags, start=2):
        for col_idx, value in enumerate(tag, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # ===== 7. BUDGETS =====
    ws = wb['Budgets']
    headers = ['Month', 'Category', 'Budgeted', 'Spent', 'Remaining', 'Progress %', 'Status']

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    budgets = [
        (datetime(2024, 1, 1), 'Housing', 1500.00),
        (datetime(2024, 1, 1), 'Food', 600.00),
        (datetime(2024, 1, 1), 'Transportation', 300.00),
        (datetime(2024, 1, 1), 'Entertainment', 200.00),
        (datetime(2024, 1, 1), 'Shopping', 400.00),
        (datetime(2024, 1, 1), 'Utilities', 200.00),
    ]

    for row_idx, (month, category, budgeted) in enumerate(budgets, start=2):
        ws.cell(row=row_idx, column=1, value=month)
        ws.cell(row=row_idx, column=2, value=category)
        ws.cell(row=row_idx, column=3, value=budgeted)
        # Spent formula
        ws.cell(row=row_idx, column=4, value=f'=SUMIFS(Transactions!C:C,Transactions!D:D,B{row_idx},Transactions!A:A,">="&DATE(YEAR(A{row_idx}),MONTH(A{row_idx}),1),Transactions!A:A,"<="&EOMONTH(A{row_idx},0),Transactions!B:B,"Expense")')
        # Remaining formula
        ws.cell(row=row_idx, column=5, value=f'=C{row_idx}-D{row_idx}')
        # Progress %
        ws.cell(row=row_idx, column=6, value=f'=D{row_idx}/C{row_idx}')
        # Status
        ws.cell(row=row_idx, column=7, value=f'=IF(F{row_idx}>1,"Over Budget",IF(F{row_idx}>0.8,"Warning","On Track"))')

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # ===== 8. GOALS =====
    ws = wb['Goals']
    headers = ['ID', 'Goal Name', 'Target Amount', 'Current Amount', 'Progress %', 'Deadline', 'Monthly Target', 'Status']

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    goals = [
        (1, 'Emergency Fund', 10000.00, 5000.00, datetime(2024, 12, 31)),
        (2, 'Vacation', 3000.00, 1200.00, datetime(2024, 6, 30)),
        (3, 'New Car', 25000.00, 8000.00, datetime(2025, 12, 31)),
        (4, 'House Down Payment', 50000.00, 15000.00, datetime(2026, 12, 31)),
    ]

    for row_idx, (id, name, target, current, deadline) in enumerate(goals, start=2):
        ws.cell(row=row_idx, column=1, value=id)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=target)
        ws.cell(row=row_idx, column=4, value=current)
        ws.cell(row=row_idx, column=5, value=f'=D{row_idx}/C{row_idx}')
        ws.cell(row=row_idx, column=6, value=deadline)
        ws.cell(row=row_idx, column=7, value=f'=IF(F{row_idx}="","N/A",(C{row_idx}-D{row_idx})/((F{row_idx}-TODAY())/30))')
        ws.cell(row=row_idx, column=8, value=f'=IF(D{row_idx}>=C{row_idx},"Achieved",IF(E{row_idx}>=0.75,"On Track",IF(E{row_idx}>=0.5,"Behind","Urgent")))')

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ===== 9. CREDIT CARDS =====
    ws = wb['Credit_Cards']
    headers = ['ID', 'Card Name', 'Bank', 'Last 4 Digits', 'Credit Limit', 'Current Balance', 'Available Credit', 'Interest Rate', 'Due Date', 'Status']

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    credit_cards = [
        (1, 'Chase Sapphire', 'Chase', '1234', 10000.00, 15, 'Active'),
        (2, 'Amex Gold', 'American Express', '5678', 15000.00, 20, 'Active'),
    ]

    for row_idx, (id, name, bank, last4, limit, due, status) in enumerate(credit_cards, start=2):
        ws.cell(row=row_idx, column=1, value=id)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=bank)
        ws.cell(row=row_idx, column=4, value=last4)
        ws.cell(row=row_idx, column=5, value=limit)
        ws.cell(row=row_idx, column=6, value=f'=SUMIFS(CC_Transactions!D:D,CC_Transactions!B:B,A{row_idx})-SUMIFS(CC_Payments!C:C,CC_Payments!B:B,A{row_idx})')
        ws.cell(row=row_idx, column=7, value=f'=E{row_idx}-F{row_idx}')
        ws.cell(row=row_idx, column=8, value='18.99%')
        ws.cell(row=row_idx, column=9, value=due)
        ws.cell(row=row_idx, column=10, value=status)

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # ===== 10. CC_TRANSACTIONS =====
    ws = wb['CC_Transactions']
    headers = ['Date', 'Card ID', 'Card Name', 'Amount', 'Category', 'Description', 'Status']

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    cc_trans = [
        (datetime(2024, 1, 15), 1, 'Chase Sapphire', 150.00, 'Dining Out', 'Restaurant', 'Posted'),
        (datetime(2024, 1, 16), 1, 'Chase Sapphire', 50.00, 'Gas', 'Gas station', 'Posted'),
        (datetime(2024, 1, 17), 2, 'Amex Gold', 1200.00, 'Shopping', 'Electronics', 'Posted'),
    ]

    for row_idx, trans in enumerate(cc_trans, start=2):
        for col_idx, value in enumerate(trans, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # ===== 11. CC_PAYMENTS =====
    ws = wb['CC_Payments']
    headers = ['Date', 'Card ID', 'Payment Amount', 'From Account', 'Payment Method', 'Status']

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    payments = [
        (datetime(2024, 1, 20), 1, 500.00, 'Chase Checking', 'Bank Transfer', 'Completed'),
        (datetime(2024, 1, 21), 2, 800.00, 'Chase Checking', 'Bank Transfer', 'Completed'),
    ]

    for row_idx, payment in enumerate(payments, start=2):
        for col_idx, value in enumerate(payment, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ===== 12. BILLS =====
    ws = wb['Bills']
    headers = ['ID', 'Bill Name', 'Payee', 'Amount', 'Due Date', 'Status', 'Category', 'Auto-Pay', 'Account']

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    bills = [
        (1, 'Electric Bill', 'Power Company', 120.00, 15, 'Pending', 'Utilities', 'No', 'Chase Checking'),
        (2, 'Internet', 'ISP Provider', 79.99, 10, 'Paid', 'Utilities', 'Yes', 'Chase Checking'),
        (3, 'Phone Bill', 'Phone Company', 65.00, 5, 'Paid', 'Utilities', 'Yes', 'Chase Checking'),
        (4, 'Insurance', 'Insurance Co', 250.00, 1, 'Pending', 'Insurance', 'Yes', 'Chase Checking'),
    ]

    for row_idx, bill in enumerate(bills, start=2):
        for col_idx, value in enumerate(bill, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # ===== 13. BILL_PAYMENTS =====
    ws = wb['Bill_Payments']
    headers = ['Date', 'Bill ID', 'Bill Name', 'Amount Paid', 'From Account', 'Payment Method', 'Status']

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    bill_payments = [
        (datetime(2024, 1, 15), 1, 'Electric Bill', 120.00, 'Chase Checking', 'Bank Transfer', 'Completed'),
        (datetime(2024, 1, 10), 2, 'Internet', 79.99, 'Chase Checking', 'Auto-Pay', 'Completed'),
    ]

    for row_idx, payment in enumerate(bill_payments, start=2):
        for col_idx, value in enumerate(payment, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # ===== 14. REPORTS =====
    ws = wb['Reports']

    ws['A1'] = 'MONTHLY REPORT'
    ws['A1'].font = Font(bold=True, size=14)
    ws['B1'] = datetime.now().strftime('%B %Y')

    ws['A3'] = 'Total Income'
    ws['B3'] = '=SUMIFS(Transactions!C:C,Transactions!B:B,"Income",Transactions!A:A,">="&EOMONTH(TODAY(),-1)+1)'
    ws['A4'] = 'Total Expenses'
    ws['B4'] = '=SUMIFS(Transactions!C:C,Transactions!B:B,"Expense",Transactions!A:A,">="&EOMONTH(TODAY(),-1)+1)'
    ws['A5'] = 'Net Savings'
    ws['B5'] = '=B3-B4'
    ws['A6'] = 'Savings Rate'
    ws['B6'] = '=B5/B3'

    ws['A8'] = 'Top 5 Expense Categories'
    ws['A8'].font = Font(bold=True)

    ws['D1'] = 'YEARLY REPORT'
    ws['D1'].font = Font(bold=True, size=14)
    ws['E1'] = str(datetime.now().year)

    ws['D3'] = 'Total Income'
    ws['E3'] = '=SUMIFS(Transactions!C:C,Transactions!B:B,"Income",Transactions!A:A,">="&DATE(YEAR(TODAY()),1,1))'
    ws['D4'] = 'Total Expenses'
    ws['E4'] = '=SUMIFS(Transactions!C:C,Transactions!B:B,"Expense",Transactions!A:A,">="&DATE(YEAR(TODAY()),1,1))'
    ws['D5'] = 'Net Savings'
    ws['E5'] = '=E3-E4'
    ws['D6'] = 'Average Monthly Income'
    ws['E6'] = '=E3/MONTH(TODAY())'
    ws['D7'] = 'Average Monthly Expense'
    ws['E7'] = '=E4/MONTH(TODAY())'

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 25

    # ===== 15. CHARTS =====
    ws = wb['Charts']

    ws['A1'] = 'VISUAL ANALYTICS'
    ws['A1'].font = Font(bold=True, size=16)

    ws['A3'] = 'Charts and visualizations can be created by:'
    ws['A4'] = '1. Selecting data from other sheets'
    ws['A5'] = '2. Insert → Chart in Google Sheets'
    ws['A6'] = '3. Choose chart type (Pie, Bar, Line, etc.)'

    ws['A8'] = 'Recommended Charts:'
    ws['A9'] = '- Income vs Expenses (Combo Chart)'
    ws['A10'] = '- Spending by Category (Pie Chart)'
    ws['A11'] = '- Net Worth Trend (Line Chart)'
    ws['A12'] = '- Budget Progress (Bar Chart)'

    ws.column_dimensions['A'].width = 50

    # Save the file
    filename = 'Daybook_Financial_Tracker.xlsx'
    wb.save(filename)
    print(f"✅ Excel file created successfully: {filename}")
    print(f"📊 Total sheets: {len(wb.sheetnames)}")
    print(f"📝 Ready to import into Google Sheets!")
    print(f"\nTo import:")
    print(f"1. Go to https://sheets.google.com")
    print(f"2. File → Import → Upload")
    print(f"3. Select {filename}")
    print(f"4. Choose 'Replace spreadsheet'")
    print(f"5. Click 'Import data'")
    print(f"\n🎉 Done! Your financial tracker is ready to use!")

if __name__ == '__main__':
    try:
        create_daybook_excel()
    except ImportError:
        print("❌ Error: openpyxl library not found")
        print("📦 Install it with: pip install openpyxl")
        print("\nFull installation:")
        print("  pip install openpyxl")
